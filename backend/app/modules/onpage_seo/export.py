"""Export engine for the On-Page SEO module.

Generates downloadable reports of page audit data in three formats:

  * **CSV** — Python stdlib (``csv``);
  * **Excel (.xlsx)** — ``openpyxl`` with a header row and autosized columns;
  * **PDF** — ``reportlab`` table output.

Files are written under ``settings.EXPORT_BASE_DIR`` (a storage subfolder
that the app already creates) and served back as download URLs.
"""
from __future__ import annotations

import csv
import logging
import uuid
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

from app.core.config import settings

logger = logging.getLogger("app.modules.onpage_seo.export")


class ExportError(Exception):
    """Raised when a requested export format cannot be produced."""


class ExportService:
    """Produces page-audit export files."""

    SUPPORTED_FORMATS = ("csv", "xlsx", "pdf")

    def __init__(self, base_dir: str | None = None) -> None:
        self.base_dir = Path(base_dir or settings.EXPORT_BASE_DIR)
        self.base_dir.mkdir(parents=True, exist_ok=True)

    # --- Public API ------------------------------------------------------

    async def export_pages(
        self,
        rows: List[Dict[str, Any]],
        fmt: str = "csv",
        scope: str = "pages",
    ) -> Dict[str, Any]:
        """Export a list of page dicts to ``fmt`` and return download info."""
        fmt = fmt.lower()
        if fmt not in self.SUPPORTED_FORMATS:
            raise ExportError(f"Unsupported export format: {fmt}")

        if not rows:
            raise ExportError("No rows to export")

        headers = list(rows[0].keys())
        data = [[_cell(row.get(h)) for h in headers] for row in rows]

        filename = f"{scope}-{uuid.uuid4().hex[:8]}.{fmt}"
        filepath = self.base_dir / filename

        if fmt == "csv":
            self._write_csv(filepath, headers, data)
        elif fmt == "xlsx":
            self._write_xlsx(filepath, headers, data)
        elif fmt == "pdf":
            self._write_pdf(filepath, headers, data)

        logger.info("Exported %d rows to %s (%s)", len(rows), filepath, fmt)
        return {
            "filename": filename,
            "format": fmt,
            "rows": len(rows),
            "download_url": f"/api/v1/exports/{filename}",
            "size_bytes": filepath.stat().st_size,
        }

    # --- Format writers ---------------------------------------------------

    @staticmethod
    def _write_csv(path: Path, headers: List[str], data: List[List[Any]]) -> None:
        with open(path, "w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerow(headers)
            writer.writerows(data)

    @staticmethod
    def _write_xlsx(path: Path, headers: List[str], data: List[List[Any]]) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter
        except ImportError as exc:  # pragma: no cover
            raise ExportError("openpyxl is not installed; cannot export xlsx") from exc

        wb = Workbook()
        ws = wb.active
        ws.title = "Pages"
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in data:
            ws.append(row)

        # Autosize columns (bounded).
        for idx, header in enumerate(headers, start=1):
            max_len = len(str(header))
            for row in data:
                max_len = max(max_len, len(str(_cell(row[idx - 1] if idx - 1 < len(row) else ""))))
            ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 60)

        wb.save(path)

    @staticmethod
    def _write_pdf(path: Path, headers: List[str], data: List[List[Any]]) -> None:
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
        except ImportError as exc:  # pragma: no cover
            raise ExportError("reportlab is not installed; cannot export pdf") from exc

        doc = SimpleDocTemplate(
            str(path),
            pagesize=landscape(A4),
            rightMargin=0.5 * inch,
            leftMargin=0.5 * inch,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch,
        )
        table_data = [headers] + data
        # Truncate very long cells so the table stays legible.
        table_data = [[(_cell(c))[:120] for c in row] for row in table_data]

        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        doc.build([table])


def _cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, dict)):
        import json

        return json.dumps(value, default=str)
    return str(value)
